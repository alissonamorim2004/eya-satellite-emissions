"""
validate_points.py -- Cruzamento de pontos NO2/CH4 com lixoes (OSM + SNIS + known_sites.json)
Vilhena, RO -- bbox expandido ~50km ao redor do centro urbano

Fontes de locais de residuos (em ordem de prioridade):
  1. OpenStreetMap (Overpass API)
  2. SNIS RS -- download CSV anual do Ministerio do Desenvolvimento Regional
  3. known_sites.json -- arquivo manual local (fallback quando OSM/SNIS nao tem dados)
"""

import math
import json
import os
import sys
import requests
from supabase import create_client

# Leitura manual do .env (evita problema de dotenv em stdin/shell)
def _load_env(path):
    if not os.path.exists(path):
        return
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            os.environ.setdefault(k.strip(), v.strip())

_load_env(os.path.join(os.path.dirname(__file__), ".env"))

# ── Configuracao ──────────────────────────────────────────────────────────────

BBOX = {
    "lat_min": -13.1,
    "lon_min": -60.5,
    "lat_max": -12.3,
    "lon_max": -59.7,
}

# Severidades a incluir (all = tudo)
SEVERIDADES = ["low", "medium", "high", "critical"]

DIST_CORRELACAO_KM = 7   # distancia maxima para correlacao forte

# Codigo IBGE do municipio monitorado (usado para filtrar known_sites.json)
IBGE_MUNICIPIO = 1100304   # Vilhena, RO


# ── Parte 1: Buscar lixoes via Overpass API ───────────────────────────────────

def buscar_lixoes_osm(lat_min, lon_min, lat_max, lon_max):
    """
    Consulta OSM para residuos solidos / lixoes dentro do bbox.
    Inclui varios tags comuns no Brasil: amenity=waste_disposal,
    landuse=landfill, waste=landfill, e tambem relation de tipo landfill.
    """
    query = f"""
    [out:json][timeout:30];
    (
      node["amenity"="waste_disposal"]({lat_min},{lon_min},{lat_max},{lon_max});
      way["amenity"="waste_disposal"]({lat_min},{lon_min},{lat_max},{lon_max});
      node["landuse"="landfill"]({lat_min},{lon_min},{lat_max},{lon_max});
      way["landuse"="landfill"]({lat_min},{lon_min},{lat_max},{lon_max});
      relation["landuse"="landfill"]({lat_min},{lon_min},{lat_max},{lon_max});
      node["waste"="landfill"]({lat_min},{lon_min},{lat_max},{lon_max});
      way["waste"="landfill"]({lat_min},{lon_min},{lat_max},{lon_max});
      node["amenity"="recycling"]({lat_min},{lon_min},{lat_max},{lon_max});
      way["man_made"="wastewater_plant"]({lat_min},{lon_min},{lat_max},{lon_max});
    );
    out center;
    """
    resp = requests.post(
        "https://overpass-api.de/api/interpreter",
        data={"data": query},
        timeout=35,
    )
    resp.raise_for_status()
    elements = resp.json().get("elements", [])
    return elements


# ── Parte 2: Calcular distancia haversine ─────────────────────────────────────

def distancia_km(lat1, lon1, lat2, lon2):
    R = 6371
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = (
        math.sin(dlat / 2) ** 2
        + math.cos(math.radians(lat1))
        * math.cos(math.radians(lat2))
        * math.sin(dlon / 2) ** 2
    )
    return R * 2 * math.asin(math.sqrt(max(0.0, a)))


# ── Parte 2b: SNIS RS -- download CSV anual ───────────────────────────────────

def buscar_lixoes_snis(ibge_municipio: int) -> list:
    """
    Tenta baixar o diagnostico anual de residuos solidos do SNIS e filtrar
    pelo codigo IBGE do municipio. Retorna lista no mesmo formato que o OSM.
    Falha graciosamente se o servidor nao estiver acessivel.
    """
    # URLs do diagnostico RS do SNIS (tentadas em ordem)
    SNIS_URLS = [
        "https://www.snis.gov.br/downloads/diagnosticos/rs/2022/RS_2022_MUNICIPIOS.xlsx",
        "http://snis.gov.br/downloads/diagnosticos/rs/2022/RS_2022_MUNICIPIOS.xlsx",
    ]

    import io
    try:
        import openpyxl
    except ImportError:
        print("      [SNIS] openpyxl nao instalado. Execute: pip install openpyxl")
        return []

    conteudo = None
    for url in SNIS_URLS:
        try:
            r = requests.get(url, timeout=20, verify=False,
                             headers={"User-Agent": "EYA-Monitor/1.0"})
            r.raise_for_status()
            conteudo = r.content
            print(f"      [SNIS] Download OK: {url[:60]} ({len(conteudo)//1024} KB)")
            break
        except Exception as e:
            print(f"      [SNIS] Falha em {url[:50]}: {type(e).__name__}")

    if not conteudo:
        return []

    wb = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

    # Colunas esperadas no SNIS RS
    col_ibge = next((i for i, h in enumerate(headers) if "IBGE" in h.upper() or "CO_MUNICIPIO" in h.upper()), None)
    col_lat  = next((i for i, h in enumerate(headers) if "LATITUDE" in h.upper()), None)
    col_lon  = next((i for i, h in enumerate(headers) if "LONGITUDE" in h.upper()), None)
    col_nome = next((i for i, h in enumerate(headers) if "NOME" in h.upper() and "MUNICIPIO" in h.upper()), None)
    col_tipo = next((i for i, h in enumerate(headers) if "TIPO" in h.upper() or "UNIDADE" in h.upper()), None)

    if col_ibge is None or col_lat is None or col_lon is None:
        print(f"      [SNIS] Colunas nao encontradas. Headers: {headers[:10]}")
        return []

    sites = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            ibge_val = int(str(row[col_ibge]).strip())
        except (TypeError, ValueError):
            continue
        if ibge_val != ibge_municipio:
            continue
        lat = row[col_lat]
        lon = row[col_lon]
        if lat is None or lon is None:
            continue
        try:
            lat, lon = float(lat), float(lon)
        except (TypeError, ValueError):
            continue
        nome = str(row[col_nome]) if col_nome is not None else "SNIS"
        tipo = str(row[col_tipo]) if col_tipo is not None else ""
        sites.append({
            "id":   f"snis_{ibge_val}",
            "type": "snis",
            "lat":  lat,
            "lon":  lon,
            "tags": {"name": nome, "tipo": tipo, "fonte": "SNIS RS 2022"},
        })

    print(f"      [SNIS] {len(sites)} registro(s) para IBGE {ibge_municipio}")
    return sites


# ── Parte 2c: Locais manuais (known_sites.json) ───────────────────────────────

def buscar_lixoes_manuais(ibge_municipio: int) -> list:
    """
    Carrega known_sites.json e filtra pelo codigo IBGE.
    Permite adicionar manualmente coordenadas de aterros nao mapeados no OSM/SNIS.
    """
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "known_sites.json")
    if not os.path.exists(path):
        return []
    try:
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        sites = [
            {
                "id":   f"manual_{s.get('ibge')}_{i}",
                "type": "manual",
                "lat":  float(s["lat"]),
                "lon":  float(s["lon"]),
                "tags": {
                    "name":   s.get("nome", "sem nome"),
                    "tipo":   s.get("tipo", ""),
                    "fonte":  s.get("fonte", "manual"),
                    "obs":    s.get("obs", ""),
                },
            }
            for i, s in enumerate(data.get("sites", []))
            if s.get("ibge") == ibge_municipio
        ]
        print(f"      [known_sites.json] {len(sites)} registro(s) para IBGE {ibge_municipio}")
        return sites
    except Exception as e:
        print(f"      [known_sites.json] Erro ao ler: {e}")
        return []


# ── Parte 3: Buscar pontos NO2 do Supabase (filtro por bbox) ─────────────────

def buscar_pontos_supabase():
    url = os.getenv("VITE_SUPABASE_URL")
    key = os.getenv("SUPABASE_SERVICE_KEY")
    if not url or not key:
        raise EnvironmentError(
            "VITE_SUPABASE_URL e SUPABASE_SERVICE_KEY precisam estar no .env"
        )
    sb = create_client(url, key)

    q = (
        sb.table("air_pollution")
        .select("id, latitude, longitude, no2_value, severity, recorded_at, source")
        .gte("latitude",  BBOX["lat_min"])
        .lte("latitude",  BBOX["lat_max"])
        .gte("longitude", BBOX["lon_min"])
        .lte("longitude", BBOX["lon_max"])
        .in_("severity", SEVERIDADES)
        .order("no2_value", desc=True)
    )
    return q.execute().data


# ── Parte 4: Cruzamento ───────────────────────────────────────────────────────

def cruzar(pontos, lixoes):
    relatorio = []

    for ponto in pontos:
        p_lat = ponto["latitude"]
        p_lon = ponto["longitude"]

        correlacoes = []
        for lixao in lixoes:
            l_lat = lixao.get("lat") or lixao.get("center", {}).get("lat")
            l_lon = lixao.get("lon") or lixao.get("center", {}).get("lon")
            if l_lat is None or l_lon is None:
                continue
            dist = distancia_km(p_lat, p_lon, l_lat, l_lon)
            tags = lixao.get("tags", {})
            correlacoes.append(
                {
                    "nome":        tags.get("name", "sem nome"),
                    "osm_id":      lixao.get("id"),
                    "osm_type":    lixao.get("type"),
                    "tags":        tags,
                    "lat":         l_lat,
                    "lon":         l_lon,
                    "distancia_km": round(dist, 2),
                    "forte":       dist < DIST_CORRELACAO_KM,
                }
            )

        correlacoes.sort(key=lambda c: c["distancia_km"])
        relatorio.append(
            {
                "ponto": {
                    "id":          ponto.get("id"),
                    "lat":         p_lat,
                    "lon":         p_lon,
                    "no2_value":   ponto.get("no2_value"),
                    "severity":    ponto.get("severity", "?"),
                    "source":      ponto.get("source"),
                    "recorded_at": ponto.get("recorded_at", "?"),
                },
                "lixoes_proximos": correlacoes,
            }
        )

    return relatorio


# ── Parte 5: Relatorio ────────────────────────────────────────────────────────

def imprimir_relatorio(relatorio):
    fortes_total = 0

    for item in relatorio:
        p = item["ponto"]
        print(
            f"\nPonto ({p['lat']:.4f}, {p['lon']:.4f})"
            f"  NO2={p['no2_value']:.1f} ppb"
            f"  [{p['severity'].upper()}]"
            f"  {p['recorded_at'][:10]}"
        )
        if not item["lixoes_proximos"]:
            print("   Nenhum lixao/residuo mapeado no OSM dentro do bbox.")
            continue
        for c in item["lixoes_proximos"]:
            tag_str = ""
            if c["tags"]:
                kv = ", ".join(f"{k}={v}" for k, v in list(c["tags"].items())[:3])
                tag_str = f"  ({kv})"
            flag = "  *** CORRELACAO FORTE ***" if c["forte"] else ""
            print(
                f"   -> [{c['osm_type']}/{c['osm_id']}] {c['nome']}"
                f"  {c['distancia_km']} km{tag_str}{flag}"
            )
            if c["forte"]:
                fortes_total += 1

    sep = "-" * 60
    print(f"\n{sep}")
    print(f"Pontos analisados          : {len(relatorio)}")
    print(f"Correlacoes fortes (<{DIST_CORRELACAO_KM}km) : {fortes_total}")
    print(sep)


# ── Main ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")

    print("=" * 60)
    print("EYA -- Validacao de Pontos x Lixoes (Vilhena, RO)")
    print("=" * 60)

    # -- 1. OpenStreetMap (Overpass) ------------------------------------------
    print("\n[1/4] OpenStreetMap (Overpass API)...")
    lixoes_osm = buscar_lixoes_osm(**BBOX)
    print(f"      {len(lixoes_osm)} elemento(s)")
    for el in lixoes_osm:
        lat  = el.get("lat") or el.get("center", {}).get("lat")
        lon  = el.get("lon") or el.get("center", {}).get("lon")
        nome = el.get("tags", {}).get("name", "sem nome")
        print(f"      [osm/{el.get('id')}] {nome}  ({lat}, {lon})")
    if not lixoes_osm:
        print("      Nenhum dado mapeado no OSM para este bbox.")

    # -- 2. SNIS RS -----------------------------------------------------------
    print("\n[2/4] SNIS Residuos Solidos (download CSV anual)...")
    import warnings
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        lixoes_snis = buscar_lixoes_snis(IBGE_MUNICIPIO)

    # -- 3. Locais manuais (known_sites.json) ---------------------------------
    print("\n[3/4] Locais manuais (known_sites.json)...")
    lixoes_manuais = buscar_lixoes_manuais(IBGE_MUNICIPIO)
    for s in lixoes_manuais:
        obs = s["tags"].get("obs", "")
        print(f"      [{s['id']}] {s['tags']['name']}  ({s['lat']}, {s['lon']})")
        if obs:
            print(f"        OBS: {obs}")

    # Mescla todas as fontes (sem duplicatas por coordenada)
    lixoes_todos = lixoes_osm + lixoes_snis + lixoes_manuais
    print(f"\n      Total de locais de residuos: {len(lixoes_todos)}")
    if not lixoes_todos:
        print("      AVISO: Nenhum local de residuos encontrado em nenhuma fonte.")
        print("      Adicione coordenadas em pipeline/known_sites.json para habilitar o cruzamento.")

    # -- 4. Pontos Supabase ---------------------------------------------------
    print("\n[4/4] Buscando pontos de poluicao no Supabase...")
    pontos = buscar_pontos_supabase()
    print(f"      {len(pontos)} ponto(s) encontrado(s) (severidades: {SEVERIDADES})")

    if not pontos:
        print("      Nenhum ponto encontrado no bbox. Encerrando.")
        sys.exit(0)

    print("\nCruzando pontos x lixoes...")
    relatorio = cruzar(pontos, lixoes_todos)
    imprimir_relatorio(relatorio)

    output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "validation_result.json")
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(relatorio, f, ensure_ascii=False, indent=2, default=str)
    print(f"\nResultado completo salvo em: {output_path}")
